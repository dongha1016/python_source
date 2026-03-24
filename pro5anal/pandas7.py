import pandas as pd 

# [샘플 데이터프레임 생성]
df = pd.DataFrame({
    '상품명': ['Mouse', 'Keyboard', 'Monitor'],
    '수량': [10, 5, 2],
    '가격': [12000, 25000, 300000]
})

# [파생 변수 생성] 
# 기존 컬럼들을 연산하여 '총금액'이라는 새로운 컬럼을 추가합니다.
df['총금액'] = df['수량'] * df['가격']

# [ExcelWriter 사용]
# engine='openpyxl'을 지정해야 단순 저장 외에 서식(색상, 폰트 등) 수정이 가능합니다.
# 'with' 문을 사용하여 작업이 끝나면 파일이 안전하게 닫히도록 합니다.
with pd.ExcelWriter('result5.xlsx', engine='openpyxl') as writer:
    # startrow=2: 0부터 시작하므로 3번째 행부터 데이터를 채웁니다. (위쪽 여백 확보)
    df.to_excel(writer, sheet_name='Report', index=False, startrow=2)

    # 직접 스타일을 입히기 위해 워크시트 객체(ws)를 가져옵니다.
    ws = writer.sheets['Report']

    # [셀 데이터 직접 입력]
    ws['A1'] = '상품 판매 보고서'  # 제목 입력

    from openpyxl.styles import Font  # 글꼴 스타일 설정 모듈

    # [제목 스타일링]
    # 폰트 굵게(bold), 크기(size)를 조절하여 제목답게 만듭니다.
    ws['A1'].font = Font(bold=True, size=14)

    # [헤더 스타일링]
    from openpyxl.styles import PatternFill, Alignment

    # 폰트 색상(Hex code): FFFFFF는 흰색입니다.
    header_font = Font(bold=True, color='FFFFFF')

    # 배경색 채우기: 4F81BD(파란색 계열)를 단색(solid)으로 채웁니다.
    header_fill = PatternFill(start_color='4F81BD', fill_type='solid')

    # ws[3]은 데이터프레임이 시작된 3번째 행(헤더 부분)을 의미합니다.
    for cell in ws[3]:
        cell.font = header_font               # 흰색 굵은 글씨 적용
        cell.fill = header_fill               # 파란 배경색 적용
        cell.alignment = Alignment(horizontal='center')  # 텍스트 가운데 정렬

    # [컬럼 너비 최적화]
    # 데이터 길이에 맞춰 열 너비를 자동으로 조절하는 로직입니다.
    for col in ws.columns:  # 모든 열(A, B, C...)을 순회
        max_length = 0
        col_letter = col[0].column_letter  # 컬럼 알파벳 이름 추출

        for cell in col:
            try:
                if cell.value:
                    # 셀 내용 중 가장 긴 텍스트의 길이를 계산
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        # 계산된 길이에 약간의 여유(+2)를 주어 엑셀 열 너비를 설정합니다.
        ws.column_dimensions[col_letter].width = max_length + 2

    # [숫자 표시 형식 설정]
    # 금액 데이터에 천 단위 콤마(,)를 추가하여 가독성을 높입니다.
    # iter_rows: 4행부터 끝까지, 2열(수량)부터 4열(총금액) 범위를 순회합니다.
    for row in ws.iter_rows(min_row=4, min_col=2, max_col=4):
        for cell in row:
            # 셀 값이 숫자일 때만 포맷을 적용합니다.
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'

    # [엑셀 테이블(Table) 기능 추가]
    # 엑셀의 '표 서식' 기능을 적용하여 필터링과 줄무늬 효과를 줍니다.
    from openpyxl.worksheet.table import Table, TableStyleInfo

    # 테이블 범위 지정 (A3부터 데이터 끝행까지)
    tab = Table(displayName="Table1", ref=f"A3:D{len(df)+3}")

    # 엑셀 기본 스타일 중 하나를 선택하고 줄무늬(RowStripes)를 활성화합니다.
    style = TableStyleInfo(
        name="TableStyleMedium9",   
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,        
        showColumnStripes=False
    )

    tab.tableStyleInfo = style  # 설정한 스타일을 테이블 객체에 연결
    ws.add_table(tab)           # 워크시트에 공식 '표'로 등록

    # [합계 행 추가 및 엑셀 함수 적용]
    total_row = len(df) + 4     # 데이터 다음 줄 위치 계산

    ws[f'A{total_row}'] = '합계'  # '합계'라는 글자 입력

    # 직접 숫자를 계산해서 넣는 대신, 엑셀 수식(=SUM)을 입력하여 나중에 엑셀에서 값이 바뀌어도 자동 계산되게 합니다.
    ws[f'D{total_row}'] = f'=SUM(D4:D{len(df)+3})'

    # [전체 정렬 마무리]
    # 4행부터 마지막 데이터까지 모든 셀의 텍스트를 가운데로 정렬합니다.
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')